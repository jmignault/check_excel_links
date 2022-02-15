#!/usr/bin/env python
import argparse
import sys
import openpyxl
import requests
from os.path import splitext

# setup and variables
# use an offset to skip header row
offset = 2
# print progress after this many rows
rows_done = 50

# define arguments and parse them
parser = argparse.ArgumentParser(description='Check links in an excel file.')
parser.add_argument('--ucol', dest='url_col', default=1, type=int,
                    help='index of column in input file containing URLs (zero-based)')
parser.add_argument('--scol', dest='status_col', default=5, type=int,
                    help='index of column in output file to write status codes (zero-based)')
parser.add_argument('--ccol', dest='content_type_col', default=6, type=int,
                    help='index of column in output file to write content type (zero-based)')
parser.add_argument('--lcol', dest='location_col', default=7, type=int,
                    help='index of column in output file to write redirect location (zero-based)')
parser.add_argument('--redir', dest='track_redirects', default=0, type=int,
                    help='write redirection chain to output file')
parser.add_argument('infile', help="Input file in Excel format")

args = parser.parse_args()

# keep an error count
errors = 0
# keep a redirect count
redirects = 0

# create the output filename based on the passed in filename
outfile = f"{splitext(args.infile)[0]}_checked.xlsx"

# open the file and get the active sheet
wb = openpyxl.load_workbook(filename=args.infile)
sheet = wb.active

# Progress reporting
print(f"Processing {sheet.max_row} rows in file {args.infile}.")
print(f"Will save records to XSLX file {outfile}.")

# Write in headers for additional columns
sheet.cell(row=1, column=args.status_col).value = 'STATUS CODE'
sheet.cell(row=1, column=args.content_type_col).value = 'CONTENT TYPE'
sheet.cell(row=1, column=args.location_col).value = 'LOCATION'

# calculate padding for processed rows
padding = len(str(sheet.max_row))

# Loop through the rows
for index, row in enumerate(sheet.iter_rows(min_row=offset)):
    # Every rows_done rows, print total rows processed and percentage done.
    if index > 1 and index % rows_done == 0:
        print(f"{index:{padding}} rows processed.", end=' ')
        print(f"({(index*100)/sheet.max_row:.0f}%)")
    try:
        # if URL, get headers: status code, mimetype
        # TODO: pass URL column as an argument rather than hard-code
        cv = row[args.url_col].value
        req = requests.get(cv, allow_redirects=True)
        the_cell = sheet.cell(row=index + offset, column=args.status_col)
        the_cell.value = str(req.status_code) + \
            ('; ' + req.reason if req.reason else '')
        try:
            # take just the mime type, drop encoding
            content_type = req.headers['content-type'].split(';')[0]
            the_cell = sheet.cell(row=index + offset,
                                  column=args.content_type_col)
            the_cell.value = content_type
        except KeyError:
            continue

        # check for redirect and follow if so
        if args.track_redirects != 0:
            if req.history:
                try:
                    the_cell = sheet.cell(row=index + offset,
                                          column=args.location_col)
                    fld = ''
                    for rh in req.history:
                        fld += rh.url + ';'
                    rq = requests.get(
                        req.history[-1].url, allow_redirects=True)
                    fld += f"{rq.status_code}"
                    the_cell.value = fld
                    redirects += 1
                except KeyError:
                    continue

    # record if there's no URL, then continue
    except requests.exceptions.MissingSchema:
        the_cell = sheet.cell(row=index + offset, column=args.status_col)
        the_cell.value = "No valid URL."
        if index > 0:
            print(f"No valid URL for line {index}.")
            errors += 1
        continue
    except requests.exceptions.ConnectionError:
        the_cell = sheet.cell(row=index + offset, column=args.status_col)
        the_cell.value = "Connection was refused."
        if index > 0:
            print(f"The connection was refused to {cv}: line {index}.")
            errors += 1
        continue

try:
    # try to save the output file
    wb.save(outfile)
except IOError as err:
    print(f"Could not save records to {outfile}: {str(err)}")
    exit(1)

# Done. print a message and exit.
print(f"Finished. {sheet.max_row} rows processed, saved to file {outfile}.", end=' ')
print(f"{errors} errors reported({(errors * 100)/sheet.max_row: .0f}%)")
print(f"{redirects} redirects reported({(redirects * 100)/sheet.max_row: .0f}%)")
